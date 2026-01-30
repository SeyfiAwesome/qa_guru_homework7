import zipfile
import csv
from zipfile import ZipFile
from openpyxl.reader.excel import load_workbook
import pytest
from pypdf import PdfWriter, PdfReader
from openpyxl import Workbook


@pytest.fixture
def zip_with_files(tmp_path):
    # csv
    csv_path = tmp_path / 'users.csv'
    with open(csv_path, 'w', newline='') as f:
        writer = csv.writer(f, delimiter=';')
        writer.writerow(['id', 'name'])
        writer.writerow(['1', 'Vlada'])
        writer.writerow(['2', 'Seyfi'])
        writer.writerow(['3', 'Petya'])

    # xlsx
    xlsx_path = tmp_path / 'products.xlsx'
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = 'product'
    sheet['A2'] = 'Laptop'
    workbook.save(xlsx_path)

    # pdf
    pdf_path = tmp_path / 'info.pdf'
    writer = PdfWriter()
    writer.add_blank_page(width=200, height=200)
    with open(pdf_path, 'wb') as f:
        writer.write(f)

    # Создание архива из csv, xlsx и pdf
    zip_path = tmp_path / 'files.zip'
    with ZipFile(zip_path, 'w') as zip_file:
        zip_file.write(csv_path, 'users.csv')
        zip_file.write(xlsx_path, 'products.xlsx')
        zip_file.write(pdf_path, 'info.pdf')

    return zip_path


def test_csv_from_zip(zip_with_files):
    with zipfile.ZipFile(zip_with_files) as zip_file:
        with zip_file.open('users.csv') as _:
            content = _.read().decode()
            rows = list(csv.DictReader(content.splitlines(), delimiter=';'))
    assert rows[1]['name'] == "Seyfi"


def test_xlsx_from_zip(zip_with_files):
    with zipfile.ZipFile(zip_with_files) as zip_file:
        with zip_file.open('products.xlsx') as _:
            wb = load_workbook(_)
            sheet = wb.active
            value = sheet["A2"].value
    assert value == "Laptop"


def test_pdf_from_zip(zip_with_files):
    with zipfile.ZipFile(zip_with_files) as zip_file:
        with zip_file.open('info.pdf') as _:
            reader = PdfReader(_)
            page = reader.pages[0]

    assert page is not None
